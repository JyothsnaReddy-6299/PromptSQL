import pandas as pd
from io import BytesIO
from datetime import datetime
from app.utils.tz_helper import get_ist_time
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def generate_pdf(question: str, summary: str, sql: str, records: list) -> bytes:
    buffer = BytesIO()
    
    cols = list(records[0].keys()) if records else []
    num_cols = len(cols)

    # 1. Page Orientation Choice: Landscape if we have many columns to display cleanly
    if num_cols > 6:
        pagesize = (792, 612) # Landscape Letter
        usable_width = 720 # 792 - 72 (margins)
    else:
        pagesize = (612, 792) # Portrait Letter
        usable_width = 540 # 612 - 72 (margins)

    doc = SimpleDocTemplate(
        buffer,
        pagesize=pagesize,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    # Custom warm terracotta + cream theme styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#C35237'), # Terracotta
        spaceAfter=10
    )
    
    h2_style = ParagraphStyle(
        'SectionHeading',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=11,
        textColor=colors.HexColor('#2A1E1C'), # Deep Chocolate
        spaceBefore=12,
        spaceAfter=5,
        borderColor=colors.HexColor('#F3ECE6'),
        borderWidth=0.5,
        borderPadding=(0, 0, 2, 0)
    )
    
    body_style = ParagraphStyle(
        'BodyText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        textColor=colors.HexColor('#352B26'),
        leading=11.5
    )
    
    sql_style = ParagraphStyle(
        'SQLText',
        parent=styles['Code'],
        fontName='Courier',
        fontSize=7.5,
        textColor=colors.HexColor('#713123'), # Terracotta dark
        backColor=colors.HexColor('#FAF8F5'),
        borderPadding=5,
        borderWidth=0.5,
        borderColor=colors.HexColor('#F3ECE6'),
        spaceAfter=8
    )

    # 2. Dynamic Font Scaling inside data cells based on column density
    if num_cols <= 5:
        cell_font_size = 9
        cell_leading = 12
    elif num_cols <= 8:
        cell_font_size = 8
        cell_leading = 11
    elif num_cols <= 12:
        cell_font_size = 6.5
        cell_leading = 9
    else:
        cell_font_size = 5.5
        cell_leading = 7.5

    cell_style = ParagraphStyle(
        'CellText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=cell_font_size,
        textColor=colors.HexColor('#352B26'),
        leading=cell_leading
    )
    
    header_style = ParagraphStyle(
        'HeaderCellText',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=cell_font_size,
        textColor=colors.HexColor('#C35237'),
        leading=cell_leading
    )
    
    # Title & Generation Meta
    story.append(Paragraph("PromptSQL Analytics Report", title_style))
    story.append(Paragraph(f"Report Generation Timestamp: {get_ist_time().strftime('%Y-%m-%d %H:%M:%S')}", body_style))
    story.append(Spacer(1, 10))
    
    # User Question
    story.append(Paragraph("User Query Question", h2_style))
    story.append(Paragraph(question, body_style))
    story.append(Spacer(1, 8))
    
    # Generated SQL
    story.append(Paragraph("Generated SQL Query", h2_style))
    formatted_sql = sql.replace('\n', '<br/>').replace(' ', '&nbsp;')
    story.append(Paragraph(formatted_sql, sql_style))
    story.append(Spacer(1, 8))
    
    # AI Summary
    story.append(Paragraph("AI Executive Summary & Trends", h2_style))
    formatted_summary = summary.replace('\n', '<br/>')
    story.append(Paragraph(formatted_summary, body_style))
    story.append(Spacer(1, 12))
    
    # Data Table
    story.append(Paragraph("Retrieved Data Results", h2_style))
    
    if not records:
        story.append(Paragraph("No records returned by database execution.", body_style))
    else:
        # Header row
        table_data = [[Paragraph(f"<b>{c}</b>", header_style) for c in cols]]
        
        # Populate all records (No truncation vertically)
        for row in records:
            row_cells = []
            for col in cols:
                val = row.get(col)
                val_str = "null" if val is None else str(val)
                row_cells.append(Paragraph(val_str, cell_style))
            table_data.append(row_cells)
            
        # Draw Table layout using dynamic column widths
        col_width = usable_width / num_cols
        t = Table(table_data, colWidths=[col_width] * num_cols)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F7ECE9')), # Light terracotta header
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#C35237')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#EFEAE4')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FAF8F5')])
        ]))
        story.append(t)
            
    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def generate_excel(records: list) -> bytes:
    buffer = BytesIO()
    df = pd.DataFrame(records)
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Query Results')
        workbook = writer.book
        worksheet = writer.sheets['Query Results']
        
        # Styles config matching Terracotta theme
        font_header = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        fill_header = PatternFill(start_color='C35237', end_color='C35237', fill_type='solid')
        align_header = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        font_cell = Font(name='Arial', size=10)
        align_cell = Alignment(horizontal='left', vertical='center')
        
        border_side = Side(border_style='thin', color='F3ECE6')
        border_cell = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        # Style headers
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_header
            cell.border = border_cell
            
        # Style values
        for row_num in range(2, len(records) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = font_cell
                cell.alignment = align_cell
                cell.border = border_cell
                
        # Auto width columns
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 11)
            
    excel_bytes = buffer.getvalue()
    buffer.close()
    return excel_bytes


def generate_csv(records: list) -> bytes:
    df = pd.DataFrame(records)
    csv_str = df.to_csv(index=False)
    return csv_str.encode('utf-8')
